from __future__ import annotations

import csv
import unicodedata
from pathlib import Path

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import Contato

FIELD_ALIASES = {
    'nome': ('nome', 'name', 'full_name', 'fullname'),
    'telefone': ('telefone', 'phone', 'phone_number', 'celular', 'whatsapp'),
    'email': ('email', 'e-mail', 'mail'),
    'ativo': ('ativo', 'active', 'enabled'),
    'opt_out': ('opt_out', 'optout', 'descadastrado', 'descadastro'),
}


def _normalize_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _normalize_key(value: str) -> str:
    normalized = unicodedata.normalize('NFKD', value or '')
    normalized = normalized.encode('ascii', 'ignore').decode('ascii')
    return ''.join(character for character in normalized.lower() if character.isalnum())


def _parse_bool(value) -> bool | None:
    text = _normalize_text(value).lower()
    if text in {'1', 'true', 't', 'sim', 's', 'yes', 'y', 'ativo', 'active'}:
        return True
    if text in {'0', 'false', 'f', 'nao', 'não', 'n', 'no', 'inativo', 'inactive'}:
        return False
    return None


def _pick_value(row: dict[str, object], aliases: tuple[str, ...]) -> str:
    for key, value in row.items():
        if _normalize_key(key) in aliases:
            return _normalize_text(value)
    return ''


def _rows_from_csv(arquivo) -> list[dict[str, object]]:
    arquivo.seek(0)
    conteudo = arquivo.read().decode('utf-8-sig')
    reader = csv.DictReader(conteudo.splitlines())
    return list(reader)


def _rows_from_xlsx(arquivo) -> list[dict[str, object]]:
    arquivo.seek(0)
    workbook = load_workbook(arquivo, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [str(value or '').strip() for value in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append({headers[index]: value for index, value in enumerate(row)})
    return data_rows


def carregar_linhas_de_arquivo(arquivo) -> list[dict[str, object]]:
    nome_arquivo = Path(getattr(arquivo, 'name', '')).name.lower()
    if nome_arquivo.endswith('.csv'):
        return _rows_from_csv(arquivo)
    if nome_arquivo.endswith(('.xlsx', '.xlsm')):
        return _rows_from_xlsx(arquivo)
    raise ValueError('Formato inválido. Use CSV ou XLSX.')


def importar_contatos_de_linhas(linhas: list[dict[str, object]]) -> dict[str, object]:
    resumo = {
        'criados': 0,
        'atualizados': 0,
        'ignorados': 0,
        'erros': [],
        'total': len(linhas),
    }

    for indice, linha in enumerate(linhas, start=2):
        nome = _pick_value(linha, FIELD_ALIASES['nome'])
        telefone = _pick_value(linha, FIELD_ALIASES['telefone'])
        email = _pick_value(linha, FIELD_ALIASES['email'])

        if not nome or not telefone or not email:
            resumo['ignorados'] += 1
            resumo['erros'].append(f'Linha {indice}: nome, telefone e email são obrigatórios.')
            continue

        ativo = _parse_bool(_pick_value(linha, FIELD_ALIASES['ativo']))
        opt_out = _parse_bool(_pick_value(linha, FIELD_ALIASES['opt_out']))

        if opt_out is True:
            ativo = False
        elif opt_out is False and ativo is None:
            ativo = True

        if ativo is None:
            ativo = True

        try:
            with transaction.atomic():
                contato_por_telefone = Contato.objects.filter(telefone=telefone).first()
                contato_por_email = Contato.objects.filter(email=email).first()

                if contato_por_telefone and contato_por_email and contato_por_telefone.pk != contato_por_email.pk:
                    resumo['ignorados'] += 1
                    resumo['erros'].append(
                        f'Linha {indice}: telefone e email já pertencem a contatos diferentes.'
                    )
                    continue

                contato = contato_por_telefone or contato_por_email
                dados = {
                    'nome': nome,
                    'telefone': telefone,
                    'email': email,
                }

                if ativo is not None:
                    dados['ativo'] = ativo

                if opt_out is not None:
                    dados['opt_out'] = opt_out
                    dados['opt_out_em'] = timezone.now() if opt_out else None
                    if opt_out:
                        dados['ativo'] = False

                if contato:
                    for campo, valor in dados.items():
                        setattr(contato, campo, valor)
                    contato.save()
                    resumo['atualizados'] += 1
                else:
                    Contato.objects.create(**dados)
                    resumo['criados'] += 1
        except Exception as exc:  # pragma: no cover - defensive summary collection
            resumo['ignorados'] += 1
            resumo['erros'].append(f'Linha {indice}: {exc}')

    return resumo


def importar_contatos(arquivo) -> dict[str, object]:
    linhas = carregar_linhas_de_arquivo(arquivo)
    return importar_contatos_de_linhas(linhas)